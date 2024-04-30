#!/usr/bin/python3
# PYTHON_ARGCOMPLETE_OK

import argparse
import argcomplete
import os
import pathlib
import json
import sys
import typing
import openpyxl.cell
import openpyxl.worksheet
import openpyxl.worksheet.worksheet
import requests
import datetime
import types
import locale
import dataclasses
import openpyxl
import copy
import enum
import math
from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError


JsonValue = typing.Union[None, int, float, str, bool, list["JsonValue"], "JsonObject"]
JsonObject = dict[str, JsonValue]
JsonList = list[JsonValue]
RequestParam = dict[str, str]

APP_NAME = "kimaiPyTool"
KIMAI_TAG_FOR_GOOGLE_CALENDAR = "gcalendar"
KIMAI_TAG_FOR_INVOICE_IN_PROGRESS = "invoiceInProgress"
GOOGLE_API_SCOPES = ['https://www.googleapis.com/auth/calendar']
GOOGLE_CLIENT_SECRET_FILE = 'google_api_secret.json'
GOOGLE_TOKEN_FILE = 'google_token.json'


T = typing.TypeVar("T")
def jsonObject2Class(cls: type[T], jsonObject: JsonObject) -> T:
    if type(jsonObject) is not dict:
        raise TypeError('Unexpected type for {}, expected {}, get {}'
                .format(jsonObject, dict, type(jsonObject)))
    d:dict[str, typing.Any] = dict()
    for fieldName, fieldType in typing.get_type_hints(cls).items():
        if fieldName not in jsonObject:
            if (typing.get_origin(fieldType) in [typing.Union, types.UnionType]
                    and isinstance(None, typing.get_args(fieldType))):
                d[fieldName] = None
                continue
            raise ValueError("{} not found in {}".format(fieldName, jsonObject))
        typeOk = False
        if typing.get_origin(fieldType) is None:
            typeOk = isinstance(jsonObject[fieldName], fieldType)
        elif typing.get_origin(fieldType) in [typing.Union, types.UnionType]:
            typeOk = isinstance(jsonObject[fieldName], typing.get_args(fieldType))
        elif typing.get_origin(fieldType) is list and len(typing.get_args(fieldType)) == 1:
            typeOk = (type(jsonObject[fieldName]) is list and
                all(isinstance(x, typing.get_args(fieldType)[0]) for x in jsonObject[fieldName])) # type: ignore
        else:
            raise NotImplementedError("The type {} is not suported yet".format(
                    typing.get_origin(fieldType)))
        if not typeOk:
            if jsonObject[fieldName] is None:
                print("The field {} cannot be None in {}".format(fieldName, jsonObject),
                      file=sys.stderr)
                sys.exit(1)
            raise TypeError('Unexpected type for field "{}" in {}, expected {}, get {}'
                    .format(fieldName, jsonObject, fieldType, type(jsonObject[fieldName])))
        d[fieldName] = jsonObject[fieldName]
    return cls(**d)


class InvoiceUnit(enum.Enum):
    HOUR = enum.auto()
    DAY = enum.auto()


class InvoiceRateRound(enum.Enum):
    SUBTOTAL = enum.auto()
    TOTAL = enum.auto()


@dataclasses.dataclass
class KimaiCustomer:
    id: int
    name: str
    number: str
    comment: None|str
    visible: bool
    billable: bool
    currency: str

    @property
    def invoiceUnit(self) -> None|InvoiceUnit:
        if self.comment is None:
            return None
        dataJson = json.loads(self.comment)
        if "invoiceUnit" in dataJson:
            return InvoiceUnit[dataJson["invoiceUnit"]]
        return None

    @property
    def invoiceUnitTranslated(self) -> None|str:
        if self.comment is None:
            return None
        dataJson = json.loads(self.comment)
        if "invoiceUnitTranslated" in dataJson:
            return dataJson["invoiceUnitTranslated"]
        return None

    @property
    def invoiceRateRound(self) -> None|InvoiceRateRound:
        if self.comment is None:
            return None
        dataJson = json.loads(self.comment)
        if "invoiceRateRound" in dataJson:
            return InvoiceRateRound[dataJson["invoiceRateRound"]]
        return None

    @property
    def invoiceRemainingHours(self) -> None|float:
        if self.comment is None:
            return None
        dataJson = json.loads(self.comment)
        if "invoiceRemainingHours" in dataJson:
            return dataJson["invoiceRemainingHours"]
        return None

    @invoiceRemainingHours.setter
    def invoiceRemainingHours(self, invoiceRemainingHours: float|None):
        dataJson = JsonObject()
        if self.comment is not None:
            dataJson = json.loads(self.comment)
        if invoiceRemainingHours is None:
            if "invoiceRemainingHours" in dataJson:
                del dataJson["invoiceRemainingHours"]
        else:
            dataJson["invoiceRemainingHours"] = invoiceRemainingHours
        self.comment = json.dumps(dataJson)

    @property
    def invoiceRemainingHoursInProgress(self) -> None|float:
        if self.comment is None:
            return None
        dataJson = json.loads(self.comment)
        if "invoiceRemainingHoursInProgress" in dataJson:
            return dataJson["invoiceRemainingHoursInProgress"]
        return None

    @invoiceRemainingHoursInProgress.setter
    def invoiceRemainingHoursInProgress(self, invoiceRemainingHoursInProgress: float|None):
        dataJson = JsonObject()
        if self.comment is not None:
            dataJson = json.loads(self.comment)
        if invoiceRemainingHoursInProgress is None:
            if "invoiceRemainingHoursInProgress" in dataJson:
                del dataJson["invoiceRemainingHoursInProgress"]
        else:
            dataJson["invoiceRemainingHoursInProgress"] = invoiceRemainingHoursInProgress
        self.comment = json.dumps(dataJson)


@dataclasses.dataclass
class KimaiCustomerDetails(KimaiCustomer):
    company: str
    vatId: None|str
    contact: None|str
    address: None|str
    country: str
    phone: None|str
    fax: None|str
    mobile: None|str
    email: None|str
    homepage: None|str
    timezone: None|str
    budget: float
    timeBudget: int


class KimaiCustomers:
    def __init__(self, jsonList: JsonList):
        self._customersById: dict[int, KimaiCustomer] = dict()
        self._idsByName: dict[str, int] = dict()
        for jsonValue in jsonList:
            if type(jsonValue) is not typing.get_origin(JsonObject):
                raise TypeError('Unexpected type for {}, expected {}, get {}'
                        .format(jsonValue, JsonObject, type(jsonValue)))
            jsonObject = typing.cast(JsonObject, jsonValue)
            customer = jsonObject2Class(KimaiCustomer, jsonObject)
            self._customersById[customer.id] = customer
            self._idsByName[customer.name] = customer.id

    @property
    def customersById(self) -> dict[int, KimaiCustomer]:
        return self._customersById

    def get(self, id: int) -> KimaiCustomer:
        return self._customersById[id]

    def getIdByName(self, name: str) -> int:
        return self._idsByName[name]


@dataclasses.dataclass
class KimaiCustomerRate:
    id: int
    rate: float
    internalRate: float|None
    isFixed: bool


class KimaiCustomerRates:
    def __init__(self, jsonList: JsonList):
        self._customerRatesById: dict[int, KimaiCustomerRate] = dict()
        for jsonValue in jsonList:
            if type(jsonValue) is not typing.get_origin(JsonObject):
                raise TypeError('Unexpected type for {}, expected {}, get {}'
                        .format(jsonValue, JsonObject, type(jsonValue)))
            jsonObject = typing.cast(JsonObject, jsonValue)
            customerRate = jsonObject2Class(KimaiCustomerRate, jsonObject)
            self._customerRatesById[customerRate.id] = customerRate

    @property
    def customerRatesById(self) -> dict[int, KimaiCustomerRate]:
        return self._customerRatesById

    def get(self, id: int) -> KimaiCustomerRate:
        return self._customerRatesById[id]


@dataclasses.dataclass
class KimaiProject:
    parentTitle: str
    customer: int
    id: int
    name: str
    start: str|None
    end: str|None
    comment: str|None
    visible: bool
    billable: bool


class KimaiProjects:
    def __init__(self, jsonList: JsonList):
        self._projectsById: dict[int, KimaiProject] = dict()
        self._idsByName: dict[str, int] = dict()
        self._idsByCustomerId: dict[int, list[int]] = dict()
        for jsonValue in jsonList:
            if type(jsonValue) is not typing.get_origin(JsonObject):
                raise TypeError('Unexpected type for {}, expected {}, get {}'
                        .format(jsonValue, JsonObject, type(jsonValue)))
            jsonObject = typing.cast(JsonObject, jsonValue)
            project = jsonObject2Class(KimaiProject, jsonObject)
            self._projectsById[project.id] = project
            if project.name in self._idsByName:
                raise ValueError('The project name "{}" already exist'.format(project.name))
            self._idsByName[project.name] = project.id
            if project.customer not in self._idsByCustomerId:
                self._idsByCustomerId[project.customer] = list()
            self._idsByCustomerId[project.customer].append(project.id)

    @property
    def projectsById(self) -> dict[int, KimaiProject]:
        return self._projectsById

    def get(self, id: int) -> KimaiProject:
        return self._projectsById[id]

    def getIdByName(self, name: str) -> int:
        return self._idsByName[name]

    def getIdsByCustomerId(self, customerId: int) -> list[int]:
        return self._idsByCustomerId[customerId]


@dataclasses.dataclass
class KimaiActivity:
    parentTitle: str
    project: int
    id: int
    name: str
    comment: str|None
    visible: bool
    billable: bool


@dataclasses.dataclass
class KimaiActivityDetails(KimaiActivity):
    budget: float
    timeBudget: int


class KimaiActivities:
    def __init__(self, jsonList: JsonList):
        self._activitiesById: dict[int, KimaiActivity] = dict()
        self._idsByName: dict[str, int] = dict()
        self._idsByProjectId: dict[int, list[int]] = dict()
        for jsonValue in jsonList:
            if type(jsonValue) is not typing.get_origin(JsonObject):
                raise TypeError('Unexpected type for {}, expected {}, get {}'
                        .format(jsonValue, JsonObject, type(jsonValue)))
            jsonObject = typing.cast(JsonObject, jsonValue)
            activity = jsonObject2Class(KimaiActivity, jsonObject)
            self._activitiesById[activity.id] = activity
            if activity.name in self._idsByName:
                raise ValueError('The activity name "{}" already exist'.format(activity.name))
            self._idsByName[activity.name] = activity.id
            if activity.project not in self._idsByProjectId:
                self._idsByProjectId[activity.project] = list()
            self._idsByProjectId[activity.project].append(activity.id)

    @property
    def activitiesById(self) -> dict[int, KimaiActivity]:
        return self._activitiesById

    def get(self, id: int) -> KimaiActivity:
        return self._activitiesById[id]

    def getIdByName(self, name: str) -> int:
        return self._idsByName[name]

    def getIdsByProjectId(self, projectId: int) -> list[int]:
        return self._idsByProjectId[projectId]


@dataclasses.dataclass
class KimaiTimeSheet:
    activity: int
    project: int
    user: int
    id: int
    begin: str
    end: str
    duration: int
    description: str
    rate: float
    internalRate: float
    exported: bool
    billable: bool
    tags: list[str]

    def getBegin(self) -> datetime.datetime:
        return datetime.datetime.fromisoformat(self.begin)

    def getEnd(self) -> datetime.datetime:
        return datetime.datetime.fromisoformat(self.end)


class KimaiTimeSheets:
    def __init__(self, jsonList: JsonList):
        self._timeSheetsById: dict[int, KimaiTimeSheet] = dict()
        for jsonValue in jsonList:
            if type(jsonValue) is not typing.get_origin(JsonObject):
                raise TypeError('Unexpected type for {}, expected {}, get {}'
                        .format(jsonValue, JsonObject, type(jsonValue)))
            jsonObject = typing.cast(JsonObject, jsonValue)
            timeSheet = jsonObject2Class(KimaiTimeSheet, jsonObject)
            self._timeSheetsById[timeSheet.id] = timeSheet

    @property
    def timesheetsById(self) -> dict[int, KimaiTimeSheet]:
        return self._timeSheetsById

    def get(self, id: int) -> KimaiTimeSheet:
        return self._timeSheetsById[id]

    def values(self):
        return self._timeSheetsById.values()


class ToDelete(enum.Enum):
    TO_DELETE = enum.auto()

class Kimai:
    def __init__(self, url: str, username:str, password: str):
        self._url = url
        self._username = username
        self._password = password

    def _buildheader(self):
        return {"X-AUTH-USER": self._username, "X-AUTH-TOKEN": self._password}

    def _runRequest(self, method:str, urlSufix: str, params:RequestParam|None={},
            data: JsonObject|None=None):
        response = requests.request(method, self._url + "/" + urlSufix, params=params,
                headers=self._buildheader(), data=data)
        if response.status_code != 200:
            print('For request "{}" get {}'.format(response.url, response.json()), file=sys.stderr)
            sys.exit(1)
        if response.headers.get("X-Total-Pages", "1") != "1":
            print('For request "{}" get too much result: {}'.format(response.url,
                    response.headers["X-Total-Count"]), file=sys.stderr)
            sys.exit(1)
        return response.json()

    def getCustomers(self) -> KimaiCustomers:
        return KimaiCustomers(self._runRequest("get", "customers"))

    def getCustomer(self, id: int) -> KimaiCustomerDetails:
        customerJson = self._runRequest("get", "customers/{}".format(id))
        return jsonObject2Class(KimaiCustomerDetails, customerJson)

    def updateCustomer(self, id: int, invoiceRemainingHours: float|None = None,
                invoiceRemainingHoursInProgress: float|None|ToDelete = None) -> KimaiCustomerDetails:
        customer = None
        data = JsonObject()
        if invoiceRemainingHours != None:
            if customer == None:
                customer = kimai.getCustomer(id)
            customer.invoiceRemainingHours = invoiceRemainingHours
            data["comment"] = customer.comment
        if invoiceRemainingHoursInProgress != None:
            if customer == None:
                customer = kimai.getCustomer(id)
            if invoiceRemainingHoursInProgress == ToDelete.TO_DELETE:
                customer.invoiceRemainingHoursInProgress = None
            else:
                customer.invoiceRemainingHoursInProgress = invoiceRemainingHoursInProgress
            data["comment"] = customer.comment
        if len(data)==0:
            print("You must set at least one argument", file=sys.stderr)
            sys.exit(1)
        customerJson = self._runRequest("patch", "customers/{}".format(id), data=data)
        return jsonObject2Class(KimaiCustomerDetails, customerJson)

    def getCustomerRates(self, id: int) -> KimaiCustomerRates:
        return KimaiCustomerRates(self._runRequest("get", "customers/{}/rates".format(id)))

    def getProjects(self) -> KimaiProjects:
        projectsJon = self._runRequest("get", "projects")
        return KimaiProjects(projectsJon)

    def getActivities(self) -> KimaiActivities:
        activitiesJson = self._runRequest("get", "activities")
        return KimaiActivities(activitiesJson)

    def getActivity(self, id: int) -> KimaiActivityDetails:
        activityJson = self._runRequest("get", "activities/{}".format(id))
        return jsonObject2Class(KimaiActivityDetails, activityJson)

    def updateActivity(self, id: int, timeBudgetHour: float|None = None) -> KimaiActivityDetails:
        data = JsonObject()
        if timeBudgetHour != None:
            data["timeBudget"] = timeBudgetHour
        if len(data)==0:
            print("You must set at least one argument", file=sys.stderr)
            sys.exit(1)
        customerJson = self._runRequest("patch", "activities/{}".format(id), data=data)
        return jsonObject2Class(KimaiActivityDetails, customerJson)

    def getTimesheets(self, begin:str|None=None, maxItem: int|None=None, billable: bool|None=None,
                exported: bool|None=None, active: bool|None=None, tags: list[str]|None=None) \
                -> KimaiTimeSheets:
        params:RequestParam = RequestParam()
        if begin is not None:
            params["begin"] = begin
        if maxItem is not None:
            params["size"] = str(maxItem)
        if billable is not None:
            params["billable"] = "1" if billable else "0"
        if exported is not None:
            params["exported"] = "1" if exported else "0"
        if active is not None:
            params["active"] = "1" if active else "0"
        if tags is not None:
            params["tags[]"] = tags # type: ignore
        timesheetsJson = self._runRequest("get", "timesheets", params=params)
        return KimaiTimeSheets(timesheetsJson)

    def addTimesheet(self, userId: int, projectId: int, activityId: int, begin: str, end: str,
            description: str) -> JsonObject:
        data = JsonObject()
        data["user"] = userId
        data["project"] = projectId
        data["activity"] = activityId
        data["begin"] = begin
        data["end"] = end
        data["description"] = description
        return self._runRequest("post", "timesheets", data=data)

    def updateTimesheet(self, timeSheetId: int, tags: list[str]|None=None,
            exported: bool|None = None) -> KimaiTimeSheet:
        data = JsonObject()
        if tags != None:
            data["tags"] = ", ".join(tags)
        if exported != None:
            data["exported"] = exported
        if len(data)==0:
            print("You must set at least one argument", file=sys.stderr)
            sys.exit(1)
        timeSheetJson = self._runRequest("patch", "timesheets/{}".format(timeSheetId), data=data)
        return jsonObject2Class(KimaiTimeSheet, timeSheetJson)


def getConfigPath(fileName: str) -> str:
    configPath = os.path.join(pathlib.Path.home(), ".config", APP_NAME, fileName)
    if not os.path.exists(os.path.dirname(configPath)):
        os.mkdir(os.path.dirname(configPath))
    return configPath


@dataclasses.dataclass
class Config:
    kimaiUrl: str|None = None
    kimaiUsername: str|None = None
    kimaiToken: str|None = None
    gCalendarEmail: str|None = None
    invoiceTemplate: str|None = None
    vatRate: float|None = None

    def toJson(self) -> JsonObject:
        toReturn: JsonObject = dict()
        if self.kimaiUrl is not None:
            toReturn["kimaiUrl"] = self.kimaiUrl
        if self.kimaiUsername is not None:
            toReturn["kimaiUsername"] = self.kimaiUsername
        if self.kimaiToken is not None:
            toReturn["kimaiToken"] = self.kimaiToken
        if self.gCalendarEmail is not None:
            toReturn["gCalendarEmail"] = self.gCalendarEmail
        if self.invoiceTemplate is not None:
            toReturn["invoiceTemplate"] = self.invoiceTemplate
        if self.vatRate is not None:
            toReturn["vatRate"] = self.vatRate
        return toReturn


def googleApiGetCredentials(secretFilePath: str, tokenFilePath: str):
    credentials = None
    if os.path.exists(tokenFilePath):
        credentials = Credentials.from_authorized_user_file(tokenFilePath, GOOGLE_API_SCOPES)
    if not credentials or not credentials.valid:
        if credentials and credentials.expired and credentials.refresh_token:
            credentials.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file(secretFilePath, GOOGLE_API_SCOPES)
            credentials = flow.run_local_server(port=0)
            with open(tokenFilePath, "w") as token:
                token.write(credentials.to_json())
    return credentials


@dataclasses.dataclass
class GCalendarEvent:
    summary: str
    start: str
    end: str
    description: str|None

    @classmethod
    def fromKimaiTimeSheet(cls, timeSheet: KimaiTimeSheet, clientName: str, projectName: str,
            activityName: str):
        return cls(" - ".join([clientName, projectName, activityName]), timeSheet.begin,
                timeSheet.end, timeSheet.description)

    def toJson(self) -> JsonObject:
        toReturn: JsonObject = dict()
        toReturn["summary"] = self.summary
        toReturn["start"] = dict()
        toReturn["start"]["dateTime"] = self.start
        toReturn["end"] = dict()
        toReturn["end"]["dateTime"] = self.end
        if self.description is not None:
            toReturn["description"] = self.description
        return toReturn


def googleApiPushEventToCalendar(event: GCalendarEvent, calendarEmail: str, service):
    try:
        event = service.events().insert(calendarId=calendarEmail, body=event.toJson()).execute()
        print('Event created: {}'.format(event.get('htmlLink')))
    except HttpError as err:
        message = str(err)
        if err.resp.get('content-type', '').startswith('application/json'):
            message = json.loads(err.content).get('error').get('errors')[0].get('message')
        print('Error while pushing event "{}" at {}: "{}"'.format(event.summary, event.start,
                message))

def importEventFile(timesheetsFilePath: str, kimaiUserId: int, kimai: Kimai):
        with open(timesheetsFilePath, 'r') as eventsFile:
            eventsData = json.loads(eventsFile.read())
        projects = kimai.getProjects()
        activities = kimai.getActivities()
        for eventData in eventsData:
            projectName = eventData["projectName"]
            projectId = projects.getIdByName(projectName)
            activityId = None
            if "activityName" in eventData:
                activityName = eventData["activityName"]
                activityId = activities.getIdByName(activityName)
            else:
                activitiesIds = activities.getIdsByProjectId(projectId)
                if len(activitiesIds) != 1:
                    print("For project {} we have no or several activities {}".format(projectName),
                            file=sys.stderr)
                    sys.exit(1)
                activityId = activitiesIds[0]
            timeSheet = kimai.addTimesheet(kimaiUserId, projectId, activityId,
                    eventData["begin"], eventData["end"], eventData["description"])
            print(timeSheet)


@dataclasses.dataclass
class CraItem:
    duration: int = 0
    description: set[str] = dataclasses.field(default_factory=set)


def kimaiToGCalendar(begin: datetime.datetime, kimai: Kimai, gCalendarEmail: str):
    beginStr = datetime.datetime.isoformat(begin)
    timeSheets = kimai.getTimesheets(begin=beginStr)
    googleCalendarService = None
    customers = kimai.getCustomers()
    projects = kimai.getProjects()
    activities = kimai.getActivities()
    for timeSheet in timeSheets.values():
        if KIMAI_TAG_FOR_GOOGLE_CALENDAR not in timeSheet.tags:
            if googleCalendarService is None:
                if not os.path.exists(getConfigPath(GOOGLE_CLIENT_SECRET_FILE)):
                    print("You pust generate a google API OAuth 2.0 token file from "
                            "https://developers.google.com/google-apps/calendar/quickstart/python#prerequisites"
                            " and copy it in {}".format(getConfigPath(GOOGLE_CLIENT_SECRET_FILE)),
                            file=sys.stderr)
                    sys.exit(1)
                googleCredentials = googleApiGetCredentials(getConfigPath(
                    GOOGLE_CLIENT_SECRET_FILE), getConfigPath(GOOGLE_TOKEN_FILE))
                googleCalendarService = build("calendar", "v3", credentials=googleCredentials)
            project = projects.get(timeSheet.project)
            googleCalendarEvent = GCalendarEvent.fromKimaiTimeSheet(timeSheet, customers.get(
                    project.customer).name, project.name, activities.get(timeSheet.activity).name)
            googleApiPushEventToCalendar(googleCalendarEvent, gCalendarEmail,
                    googleCalendarService)
            tags = timeSheet.tags
            tags.append(KIMAI_TAG_FOR_GOOGLE_CALENDAR)
            kimai.addTimesheetTag(timeSheet.id, tags)


def generateCraFiles(begin: datetime.datetime, kimai: Kimai):
    beginStr = datetime.datetime.isoformat(begin)
    end = datetime.date.today()
    timeSheets = kimai.getTimesheets(begin=beginStr, maxItem=100)
    customers = kimai.getCustomers()
    projects = kimai.getProjects()
    activities = kimai.getActivities()
    craByCustomerDateProjectActivity: dict[str, dict[datetime.date, dict[str, dict[str, CraItem]]]] \
            = dict()
    locale.setlocale(locale.LC_ALL, locale.getlocale())
    dateFormat = locale.nl_langinfo(locale.D_FMT)
    activitiesByCustomerProject: dict[str, dict[str, set[str]]] = dict()
    for timeSheet in timeSheets.values():
        # Get time sheet data
        date = datetime.datetime.fromisoformat(timeSheet.begin).date()
        project = projects.get(timeSheet.project)
        customerName = customers.get(project.customer).name
        activityName = activities.get(timeSheet.activity).name
        # Add to CRA
        if customerName not in craByCustomerDateProjectActivity:
            craByCustomerDateProjectActivity[customerName] = dict()
        if date not in craByCustomerDateProjectActivity[customerName]:
            craByCustomerDateProjectActivity[customerName][date] = dict()
        if project.name not in craByCustomerDateProjectActivity[customerName][date]:
            craByCustomerDateProjectActivity[customerName][date][project.name] = dict()
        if activityName not in craByCustomerDateProjectActivity[customerName][date][project.name]:
            craByCustomerDateProjectActivity[customerName][date][project.name][activityName] \
                    = CraItem()
        craByCustomerDateProjectActivity[customerName][date][project.name][activityName].duration \
                += timeSheet.duration
        for descriptionLine in timeSheet.description.replace("\r", "").split("\n"):
            craByCustomerDateProjectActivity[customerName][date][project.name][activityName] \
                    .description.add(descriptionLine)
        # Add to activitiesByCustomerProject
        if customerName not in activitiesByCustomerProject:
            activitiesByCustomerProject[customerName] = dict()
        if project.name not in activitiesByCustomerProject[customerName]:
            activitiesByCustomerProject[customerName][project.name] = set()
        activitiesByCustomerProject[customerName][project.name].add(activityName)
    # Write file by customer
    for customerName, craByDateProjectActivity in craByCustomerDateProjectActivity.items():
        with open("{}_CRA_{}.tsv".format(end.strftime("%Y-%m"), customerName), "w") as craFile:
            # write header
            headerLine1 = "\t"
            headerLine2 = "Date\tTotal (hour)"
            for projectName, activities in activitiesByCustomerProject[customerName].items():
                for activityName in activities:
                    headerLine1 += "\t"+projectName
                    headerLine2 += "\t"+activityName
            headerLine2 += "\tDescription"
            craFile.write(headerLine1+"\n")
            craFile.write(headerLine2+"\n")
            # Write a line by date
            for date, craByProjectActivity in sorted(craByDateProjectActivity.items()):
                durationSum = 0.0
                dataLine = ""
                description: set[str] = set()
                for projectName, activities in activitiesByCustomerProject[customerName].items():
                    for activityName in activities:
                        craItem = craByProjectActivity.get(projectName, dict()).get(activityName,
                                CraItem())
                        durationSum += craItem.duration
                        description.update(craItem.description)
                        dataLine += "\t"
                        if craItem.duration != 0:
                            dataLine += locale.str(craItem.duration/3600)
                craFile.write("{}\t{}{}\t{}\n".format(date.strftime(dateFormat),
                        locale.str(durationSum/3600), dataLine, ", ".join(description).replace(
                        "\t", " ")))


@dataclasses.dataclass
class InvoiceLine:
    projectName: str
    activityName: str
    begin: datetime.date
    end: datetime.date
    rateHour: float
    durationHour: float = 0.0
    durationHourFloor: float = 0.0
    unit: InvoiceUnit|None = None
    rateRound: InvoiceRateRound|None = None
    vatRate: float = 0.0

    @property
    def unitRate(self) -> float:
        if self.unit is None:
            print('Invoice unit not define', file=sys.stderr)
            sys.exit(1)
        if self.unit == InvoiceUnit.DAY:
            return 7.0
        elif self.unit == InvoiceUnit.HOUR:
            return 1.0
        else:
            raise NotImplemented("conversion from unit {} not implemented".format(self.unit))

    @property
    def durationDay(self) -> float:
        return self.durationHour / 7.0

    @property
    def duration(self) -> float:
        return self.durationHour / self.unitRate

    @property
    def durationFloor(self) -> float:
        return self.durationHourFloor / self.unitRate

    @property
    def rateDay(self) -> float:
        return self.rateHour * 7.0

    @property
    def rate(self) -> float:
        toReturn = self.rateHour * self.unitRate
        if self.rateRound == InvoiceRateRound.SUBTOTAL:
            return round(toReturn)
        elif self.rateRound == InvoiceRateRound.TOTAL:
            return round(toReturn*(1+self.vatRate))/(1+self.vatRate)
        else:
            return toReturn

    @property
    def subtotal(self) -> float:
        return self.rate * self.duration

    @property
    def subtotalFloor(self) -> float:
        return self.rate * self.durationFloor

    def updateDurationFloor(self, value: float):
        duration = math.floor(self.duration / value) * value
        diffHour = (self.duration - duration) * self.unitRate
        self.durationHourFloor = self.durationHour - diffHour

    def __str__(self) -> str:
        return "InvoiceLine(project={}, activity={}, rate={}, duration={}, unit={}, subtotal={})" \
                .format(self.projectName, self.activityName, self.rate, self.durationFloor,
                self.unit, self.subtotalFloor)


@dataclasses.dataclass
class InvoiceHeader:
    num: int
    date: datetime.date
    subtotal: float
    subtotalFloor: float
    vatRate: float

    @property
    def id(self) -> str:
        return "F{:%Y%m}{:02}".format(self.date, self.num)

    @property
    def tax(self) -> float:
        return self.subtotal * self.vatRate

    @property
    def taxFloor(self) -> float:
        return self.subtotalFloor * self.vatRate

    @property
    def total(self) -> float:
        return self.subtotal + self.taxFloor

    @property
    def totalFloor(self) -> float:
        return self.subtotalFloor + self.taxFloor

    @property
    def vatPercent(self) -> float:
        return self.vatRate * 100.0


class Invoice:
    REMAINING_FLOOR = 0.5

    def __init__(self, num: int, customer: KimaiCustomerDetails, date: datetime.date,
            lineByProjectActivity : dict[str, dict[str, InvoiceLine]], vatRate: float):
        self._num = num
        self._customer = customer
        self._date = date
        self._vatRate = vatRate
        self._lines: list[InvoiceLine] = []
        self._subtotal = 0.0
        self._subtotalFloor = 0.0
        self._remainingHour = (0.0 if customer.invoiceRemainingHours is None
                else customer.invoiceRemainingHours)
        for lineByActivity in lineByProjectActivity.values():
            for line in lineByActivity.values():
                line.unit = self._customer.invoiceUnit
                line.rateRound = self._customer.invoiceRateRound
                line.vatRate = vatRate
                line.updateDurationFloor(Invoice.REMAINING_FLOOR)
                self._remainingHour += line.durationHour - line.durationHourFloor
                REMAINING_FLOOR_HOUR = Invoice.REMAINING_FLOOR * line.unitRate
                if self._remainingHour > REMAINING_FLOOR_HOUR:
                    self._remainingHour -= REMAINING_FLOOR_HOUR
                    line.durationHourFloor += REMAINING_FLOOR_HOUR
                self._lines.append(line)
                self._subtotal += line.subtotal
                self._subtotalFloor += line.subtotalFloor
        self._remainingHour = round(self._remainingHour, 2)
        self._lines.sort(key=lambda item : item.begin)

    @property
    def header(self) -> InvoiceHeader:
        return InvoiceHeader(self._num, self._date, self._subtotal, self._subtotalFloor,
                self._vatRate)

    @property
    def remainingHour(self) -> float:
        return self._remainingHour

    def __str__(self) -> str:
        toReturn = "Invoice(customer={},\n".format(self._customer.name)
        for line in self._lines:
            toReturn += "  " + str(line) + ",\n"
        header = self.header
        toReturn += "  subtotalFloor={}, taxFloor={}, totalFloor={}, remainingHour={})".format(
                self._subtotalFloor, header.taxFloor, header.totalFloor, self._remainingHour)
        return toReturn

    def generateInvoiceFile(self, templateFilePath: str) -> float:
        templateFile = openpyxl.open(templateFilePath)
        tempateSheet = typing.cast(openpyxl.worksheet.worksheet.Worksheet, templateFile.active)
        lineIndex = 0
        isInvoiceLineCopy = False
        for rowIndex in range(1, tempateSheet.max_row+1):
            isRowContainsInvoiceLine = False
            for columnIndex in range(1, tempateSheet.max_column+1):
                cell = tempateSheet.cell(rowIndex, columnIndex)
                if type(cell.value) is str:
                    result, isInvoiceLine = self._templateReplace(cell.value, lineIndex)
                    if isInvoiceLine:
                        isRowContainsInvoiceLine = True
                        if not isInvoiceLineCopy:
                            for _ in range(len(self._lines)-1):
                                self._copyRow(tempateSheet, rowIndex, rowIndex+1)
                            isInvoiceLineCopy = True
                    if result != None:
                        cell.value = result
            if isRowContainsInvoiceLine:
                lineIndex += 1
        templateFile.save("{:%Y-%m}_facture_{}.xlsx".format(self._date, self._customer.name))
        return self._remainingHour

    def _copyRow(self, sheet: openpyxl.worksheet.worksheet.Worksheet, rowIndexSrc: int,
            rowIndexDst: int):
        sheet.insert_rows(rowIndexDst)
        if rowIndexDst < rowIndexSrc:
            rowIndexSrc += 1
        for columnIndex in range(1, sheet.max_column+1):
            cellSrc = sheet.cell(rowIndexSrc, columnIndex)
            cellDst = sheet.cell(rowIndexDst, columnIndex)
            cellDst.value = cellSrc.value
            cellDst.font = copy.copy(cellSrc.font) # type: ignore
            cellDst.alignment = copy.copy(cellSrc.alignment) # type: ignore
            cellDst.border = copy.copy(cellSrc.border) # type: ignore
            cellDst.fill = copy.copy(cellSrc.fill) # type: ignore
            cellDst.number_format = copy.copy(cellSrc.number_format) # type: ignore

    def _templateReplace(self, source: str, lineIndex: int) -> tuple[None|str|int|float, bool]:
        toReplaceIndexBegin = source.find("${")
        if toReplaceIndexBegin == -1:
            return None, False
        lastCopyIndex = 0
        result = ""
        locale.setlocale(locale.LC_ALL, locale.getlocale())
        dateFormat = locale.nl_langinfo(locale.D_FMT)
        isLine = False
        while toReplaceIndexBegin != -1:
            toReplaceIndexEnd = source.find("}", toReplaceIndexBegin)
            toReplaceName = source[toReplaceIndexBegin+2:toReplaceIndexEnd].split(".")
            toReplaceValue = None
            if toReplaceName[0] == "Customer":
                toReplaceValue = getattr(self._customer, toReplaceName[1])
            elif toReplaceName[0] == "Invoice":
                toReplaceValue = getattr(self.header, toReplaceName[1])
            elif toReplaceName[0] == "InvoiceLine":
                toReplaceValue = getattr(self._lines[lineIndex], toReplaceName[1])
                isLine = True
            else:
                print('Template value "{}" not supported'.format(toReplaceName), file=sys.stderr)
                sys.exit(1)
            if len(toReplaceName) == 3:
                if type(toReplaceValue) == datetime.date:
                    if toReplaceName[2] == "day":
                        toReplaceValue = toReplaceValue.day
                    elif toReplaceName[2] == "month":
                        toReplaceValue = toReplaceValue.month
                    elif toReplaceName[2] == "year":
                        toReplaceValue = toReplaceValue.year
                    else:
                        print('Template value "{}" not supported for a date'.format(toReplaceName),
                                file=sys.stderr)
                        sys.exit(1)
                else:
                    print('Template value "{}" not supported for {}'.format(toReplaceName,
                            type(toReplaceValue)), file=sys.stderr)
                    sys.exit(1)
            if type(toReplaceValue) == datetime.date:
                toReplaceValue = toReplaceValue.strftime(dateFormat)
            if toReplaceIndexBegin==0 and toReplaceIndexEnd == len(source)-1:
                return toReplaceValue, isLine
            # else
            result += source[lastCopyIndex:toReplaceIndexBegin]
            lastCopyIndex = toReplaceIndexEnd+1
            result += str(toReplaceValue)
            toReplaceIndexBegin = source.find("${", toReplaceIndexEnd)
        result += source[lastCopyIndex:]
        return result, isLine


def generateInvoiceFiles(kimai: Kimai, templateFilePath: str, vatRate: float):
    if not templateFilePath.endswith(".xlsx"):
        print("Only support excel xlsx file", file=sys.stderr)
        sys.exit(1)
    timeSheets = kimai.getTimesheets(maxItem=100, billable=True, exported=False, active=False)
    projects = kimai.getProjects()
    activities = kimai.getActivities()
    invoiceLineByCustomerProjectActivity: dict[int, dict[str, dict[str, InvoiceLine]]] = dict()
    for timeSheet in timeSheets.values():
        # Get time sheet data
        if KIMAI_TAG_FOR_INVOICE_IN_PROGRESS in timeSheet.tags:
            print("There are some invoice in progress {}. You must cancel or summit it before "
                  "start a new invoice".format(timeSheet), file=sys.stderr)
            sys.exit(1)
        project = projects.get(timeSheet.project)
        activity = activities.get(timeSheet.activity)
        if project.customer not in invoiceLineByCustomerProjectActivity:
            invoiceLineByCustomerProjectActivity[project.customer] = dict()
        if project.name not in invoiceLineByCustomerProjectActivity[project.customer]:
            invoiceLineByCustomerProjectActivity[project.customer][project.name] = dict()
        if activity.name not in invoiceLineByCustomerProjectActivity[project.customer][project.name]:
            rates = kimai.getCustomerRates(project.customer)
            if len(rates.customerRatesById) != 1:
                print("Too much or none rate for customer {}".format(project.customer), file=sys.stderr)
                sys.exit(1)
            rate = next(iter(rates.customerRatesById.values())).rate
            invoiceLineByCustomerProjectActivity[project.customer][project.name][activity.name] \
                    = InvoiceLine(project.name, activity.name, timeSheet.getBegin().date(),
                    timeSheet.getEnd().date(), rate)
        else:
            if timeSheet.getBegin().date() < invoiceLineByCustomerProjectActivity \
                    [project.customer][project.name][activity.name].begin:
                invoiceLineByCustomerProjectActivity[project.customer][project.name] \
                        [activity.name].begin = timeSheet.getBegin().date()
            if timeSheet.getEnd().date() > invoiceLineByCustomerProjectActivity[project.customer] \
                    [project.name][activity.name].end:
                invoiceLineByCustomerProjectActivity[project.customer][project.name] \
                        [activity.name].end = timeSheet.getEnd().date()
        invoiceLineByCustomerProjectActivity[project.customer][project.name][activity.name] \
                .durationHour += timeSheet.duration/3600.0
        invoiceLine = invoiceLineByCustomerProjectActivity[project.customer][project.name] \
                [activity.name]
        if abs(invoiceLine.rateHour * timeSheet.duration/3600.0 - timeSheet.rate) >= 0.01:
            print("Computed time sheet price not the same as kimai time sheet price: {} * {} = {} "
                    "!= {}".format(invoiceLine.rateHour, timeSheet.duration/3600.0,
                    invoiceLine.rateHour * timeSheet.duration/3600.0, timeSheet.rate),
                    file=sys.stderr)
            sys.exit(1)
    invoiceNum = 0
    for customerId, invoiceLineByProjectActivity in invoiceLineByCustomerProjectActivity.items():
        invoiceNum += 1
        customer = kimai.getCustomer(customerId)
        invoice = Invoice(invoiceNum, customer, datetime.date.today(), invoiceLineByProjectActivity,
                vatRate)
        print(invoice)
        invoice.generateInvoiceFile(templateFilePath)
        if customer.invoiceRemainingHoursInProgress is not None:
            print('The customer {} already have invoice remaining hours in progress with value {}'
                    .format(customer.name, customer.invoiceRemainingHoursInProgress),
                    file=sys.stderr)
            sys.exit(1)
        else:
            kimai.updateCustomer(customer.id, invoiceRemainingHoursInProgress=invoice.remainingHour)
            print("Customer: {} invoiceRemainingHoursInProgress: {} => {}".format(customer.name,
                    customer.invoiceRemainingHoursInProgress, invoice.remainingHour))
            customer.invoiceRemainingHoursInProgress = invoice.remainingHour
    for timeSheet in timeSheets.values():
        tags = timeSheet.tags.copy()
        tags.append(KIMAI_TAG_FOR_INVOICE_IN_PROGRESS)
        kimai.updateTimesheet(timeSheet.id, tags=tags)
        print("TimeSheet: {} tags: {} => {}".format(timeSheet.begin, timeSheet.tags, tags))
        timeSheet.tags = tags


if __name__ == '__main__':
    parser = argparse.ArgumentParser(description="Kimai cli tool")
    groupAction = parser.add_mutually_exclusive_group(required=True)
    groupAction.add_argument('--configure', action='store_true', help="Save params in config file")
    groupAction.add_argument('--getCustomers', action='store_true',
            help="Display a json list of customers")
    groupAction.add_argument('--getCustomer', type=int,
            help="Display a json object of the given customer")
    groupAction.add_argument('--updateCustomer', type=int,
            help="Update the given customer")
    groupAction.add_argument('--getCustomerRate', type=int,
            help="Display a json list of rate associate with the given customer")
    groupAction.add_argument('--getProjects', action='store_true',
            help="Display a json list of projects")
    groupAction.add_argument('--getActivities', action='store_true',
            help="Display a json list of activities")
    groupAction.add_argument('--getActivity', type=int,
            help="Display a json object of the given activity")
    groupAction.add_argument('--updateActivity', type=int,
            help="Update the given activity")
    groupAction.add_argument('--getTimesheets', action='store_true',
            help="Display a json list of timesheets")
    groupAction.add_argument('--setTimesheets', type=str, help="Import events from file")
    groupAction.add_argument('--toGCalendar', type=lambda s: datetime.datetime.strptime(s,
            '%Y-%m-%d'), help="Copy events from the given date in format YYYY-MM-DD not tag {} to "
            "Google calendar".format(KIMAI_TAG_FOR_GOOGLE_CALENDAR))
    groupAction.add_argument('--cra', type=lambda s: datetime.datetime.strptime(s,
            '%Y-%m-%d'), help="Generate a file in current dir for each customer with one line by "
            "day, project and activity with duration in hour and description")
    groupAction.add_argument('--invoice', action='store_true',
            help="Generate an invoice file in current dir for each customer using remaining hours")
    groupAction.add_argument('--invoiceInProgressCancel', action='store_true',
            help="Delete all customer invoice remaining hours in progress and remove all {} time "
            "sheet tags".format(KIMAI_TAG_FOR_INVOICE_IN_PROGRESS))
    groupAction.add_argument('--invoiceInProgressSubmit', action='store_true',
            help="Replace all customer invoice remaining hours by the in progress value and "
            "replace all {} time sheet tags by exported mark")
    parser.add_argument("--kimaiUrl", type=str, help="The Kimai url with protocol and /api "
            "exemple http://nas.local:8001/api (can be saved in config file)")
    parser.add_argument("--kimaiUsername", type=str, help="The Kimai username (can be saved in "
            "config file)")
    parser.add_argument("--kimaiToken", type=str, help="The Kimai API token (can be saved in config"
            "file)")
    parser.add_argument("--gCalendarEmail", type=str, help="The email address of the google "
            "calendar (can be saved in config file)")
    parser.add_argument("--kimaiUserId", type=int, help="The user identifier to set timesheets")
    parser.add_argument("--invoiceTemplate", type=str, help="The invoice template filepath to "
            "generate invoice file")
    parser.add_argument("--vatRate", type=float, help="The VAT rate to generate invoice")
    parser.add_argument("--timeBudget", type=float, help="When use --updateActivity update the "
            "time budget with the given time in hour")
    parser.add_argument("--invoiceRemainingHours", type=float, help="When use --updateCustomer "
            "update the invoice remaining hours with the given float")
    argcomplete.autocomplete(parser)
    args = parser.parse_args()

    configData = Config()
    configPath = getConfigPath(APP_NAME + ".json")
    if os.path.exists(configPath):
        with open(configPath, 'r') as configFile:
            configJson = json.loads(configFile.read())
            configData = jsonObject2Class(Config, configJson)
    if args.kimaiUrl:
        configData.kimaiUrl = args.kimaiUrl
    if args.kimaiUsername:
        configData.kimaiUsername = args.kimaiUsername
    if args.kimaiToken:
        configData.kimaiToken = args.kimaiToken
    if args.gCalendarEmail:
        configData.gCalendarEmail = args.gCalendarEmail
    if args.invoiceTemplate:
        configData.invoiceTemplate = args.invoiceTemplate
    if args.vatRate:
        configData.vatRate = args.vatRate

    if args.configure:
        with open(configPath, 'w') as configFile:
            json.dump(configData.toJson(), configFile, indent=4)
            configFile.write('\n')
        os.chmod(configPath, 0o600)
        sys.exit(0)
    
    if configData.kimaiUrl == None:
        print("kimai URL is not defined", file=sys.stderr)
        sys.exit(1)
    if configData.kimaiUsername == None:
        print("kimai username is not defined", file=sys.stderr)
        sys.exit(1)
    if configData.kimaiToken == None:
        print("kimai password is not defined", file=sys.stderr)
        sys.exit(1)
    kimai = Kimai(configData.kimaiUrl, configData.kimaiUsername, configData.kimaiToken)

    if args.getCustomers:
        for customer in kimai.getCustomers().customersById.values():
            print(customer)

    if args.getCustomer:
        print(kimai.getCustomer(args.getCustomer))

    if args.updateCustomer:
        print(kimai.updateCustomer(args.updateCustomer,
                invoiceRemainingHours=args.invoiceRemainingHours))

    if args.getCustomerRate:
        for customerRate in kimai.getCustomerRates(args.getCustomerRate).customerRatesById.values():
            print(customerRate)

    if args.getProjects:
        for project in kimai.getProjects().projectsById.values():
            print(project)

    if args.getActivities:
        for activity in kimai.getActivities().activitiesById.values():
            print(activity)

    if args.getActivity:
        print(kimai.getActivity(args.getActivity))

    if args.updateActivity:
        print(kimai.updateActivity(args.updateActivity, timeBudgetHour=args.timeBudget))

    if args.getTimesheets:
        for timesheet in kimai.getTimesheets().timesheetsById.values():
            print(timesheet)

    if args.setTimesheets:
        if not args.kimaiUserId:
            print("You must define the kimai user id to import data with console argument",
                    file=sys.stderr)
            sys.exit(1)
        importEventFile(args.setTimesheets, args.kimaiUserId, kimai)

    if args.toGCalendar:
        if configData.gCalendarEmail is None:
            print("Google calendar email addess not defined", file=sys.stderr)
            sys.exit(1)
        kimaiToGCalendar(args.toGCalendar, kimai, configData.gCalendarEmail)

    if args.cra:
        generateCraFiles(args.cra, kimai)

    if args.invoice:
        if configData.invoiceTemplate is None:
            print("Invoice template file path not defined", file=sys.stderr)
            sys.exit(1)
        if configData.vatRate is None:
            print("Invoice VAT rate not defined", file=sys.stderr)
            sys.exit(1)
        generateInvoiceFiles(kimai, configData.invoiceTemplate, configData.vatRate)

    if args.invoiceInProgressCancel:
        customers = kimai.getCustomers()
        for customer in customers.customersById.values():
            if customer.invoiceRemainingHoursInProgress is not None:
                kimai.updateCustomer(customer.id,
                        invoiceRemainingHoursInProgress=ToDelete.TO_DELETE)
                print("Customer: {} tags: {} => {}".format(customer.name,
                        customer.invoiceRemainingHoursInProgress, None))
                customer.invoiceRemainingHoursInProgress = None
        timesheets = kimai.getTimesheets(tags=[KIMAI_TAG_FOR_INVOICE_IN_PROGRESS])
        for timesheet in timesheets.values():
            tags = timesheet.tags.copy()
            tags.remove(KIMAI_TAG_FOR_INVOICE_IN_PROGRESS)
            kimai.updateTimesheet(timesheet.id, tags=tags)
            print("TimeSheet: {} tags: {} => {}".format(timesheet.begin, timesheet.tags, tags))
            timesheet.tags = tags

    if args.invoiceInProgressSubmit:
        timesheets = kimai.getTimesheets(tags=[KIMAI_TAG_FOR_INVOICE_IN_PROGRESS])
        for timesheet in timesheets.values():
            if timesheet.exported:
                print("The timesheet {} is tag {} and is already mark exported. Remove the tag or "
                        "the exported mark and re-run this script".format(timesheet,
                        KIMAI_TAG_FOR_INVOICE_IN_PROGRESS), file=sys.stderr)
                sys.exit(1)
            tags = timesheet.tags.copy()
            tags.remove(KIMAI_TAG_FOR_INVOICE_IN_PROGRESS)
            kimai.updateTimesheet(timesheet.id, exported=True, tags=tags)
            print("TimeSheet: {} exported: {} => {}".format(timesheet.begin, timesheet.exported,
                    True))
            print("TimeSheet: {} tags: {} => {}".format(timesheet.begin, timesheet.tags, tags))
            timesheet.exported = True
            timesheet.tags = tags
        customers = kimai.getCustomers()
        for customer in customers.customersById.values():
            if customer.invoiceRemainingHoursInProgress is not None:
                kimai.updateCustomer(customer.id,
                        invoiceRemainingHours=customer.invoiceRemainingHoursInProgress,
                        invoiceRemainingHoursInProgress=ToDelete.TO_DELETE)
                print("Customer: {} invoiceRemainingHours: {} => {}".format(customer.name,
                        customer.invoiceRemainingHours, customer.invoiceRemainingHoursInProgress))
                print("Customer: {} invoiceRemainingHoursInProgress: {} => {}".format(customer.name,
                        customer.invoiceRemainingHoursInProgress, None))
                customer.invoiceRemainingHours = customer.invoiceRemainingHoursInProgress
                customer.invoiceRemainingHoursInProgress = None
